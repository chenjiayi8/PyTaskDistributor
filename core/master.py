#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Jul 18 10:05:04 2020

@author: frank
"""

from PyTaskDistributor.util.json import (
        readJSON_to_df, writeJSON_from_df, readJSON_to_dict)

import os
import sys
from datetime import datetime
import dateutil
import pandas as pd
import numpy as np
from itertools import product as prod
#from Helper import Helper
from openpyxl import load_workbook
#import pickle
import math
import random
import uuid
import traceback
import time
#import json
import shutil

def getUUID(origin):
    rd = random.Random()
#    rd.seed(0)
    uuid_str = str(uuid.UUID(int=rd.getrandbits(128)))
    return uuid_str[:5]

class Master:
    def __init__(self, setup):
        self.setup = setup
        self.defaultFolder = os.getcwd()
        self.serverFolder = os.path.join(self.defaultFolder, 'Servers')
        self.mainFolder = self.setup['order']

        self.newTaskFolder = os.path.join(self.mainFolder, 'NewTasks')
        self.finishedTaskFolder = os.path.join(self.mainFolder, 'FinishedTasks')
        self.taskFilePath = os.path.join(self.mainFolder, 'TaskList.xlsx')
        self.lastModifiedTime = ''
        pass
    
    def main(self):
        numMin = random.randint(1,5)
        try:
            if self.lastModifiedTime != self.getFileLastModifiedTime():
                self.generateTasks()
                
            self.workloadBalance()
            lastModifiedTimeStr = datetime.strftime(self.lastModifiedTime, 
                                                      "%H:%M:%S %d/%m/%Y")
            nowTimeStr = datetime.strftime(datetime.now(),  "%H:%M:%S %d/%m/%Y")
            msg = "{}: Last task is assigned at {}, sleeping for {} mins".format(nowTimeStr, lastModifiedTimeStr, numMin)
#            print(msg)
            print("\r", msg, end='')
            time.sleep(numMin*60)
#            Helper.sleep(numMin)
            needAssistance = False
        except (KeyboardInterrupt, SystemExit):
            raise
        except:
            print ("Need assisstance for unexpected error:\n {}".format(sys.exc_info()))
            traceBackObj = sys.exc_info()[2]
            traceback.print_tb(traceBackObj)
            needAssistance = True
        return needAssistance
            
    def getFileLastModifiedTime(self):
        taskTable_modifiedTime = datetime.fromtimestamp(os.path.getmtime(self.taskFilePath))
        return taskTable_modifiedTime
    
    def emptySheetExcludeHeaders(self, ws):
#        for row in ws.columns:
        for i, row in enumerate(ws.iter_rows()):     
            if i != 0:                      
                for cell in row:
                    cell.value = None
    
    def updateServerList(self, timeout_mins=30):
        serverList = [file for file in os.listdir(self.serverFolder) if file.endswith(".json")]
        self.serverList = []
        for s in serverList:
            s_path = os.path.join(self.serverFolder, s)
            s_json = readJSON_to_dict(s_path)
            s_time = dateutil.parser.parse(s_json['updated_time'])
            diff_min = (datetime.now() - s_time).seconds/60
            if diff_min < timeout_mins:#only use if updated within N mins
                self.serverList.append(s_json)
                
    def workloadBalance(self):
        self.updateServerList()
        taskList = [file for file in os.listdir(self.newTaskFolder) if file.endswith(".json")]
        task = random.choice(taskList)#Only balance one task per time
        task_path = os.path.join(self.newTaskFolder, task)
        df = readJSON_to_df(task_path)
        df = df.sort_values('Num')
        temp = list(zip(['Task']*len(df), df['Num'].apply(str) , df['UUID']))
        index = ['-'.join(t) for t in temp]
        df.index = index
        df = df.fillna('')
        for server in self.serverList:
            if int(server['num_matlab']) == 0:
                num_target = 1
            else:
                cpu_avaiable = float(server['CPU_max']) - float(server['CPU_total'])
                cpu_per_task = float(server['CPU_matlab']) / float(server['num_matlab'])
                num_cpu = math.floor(cpu_avaiable/cpu_per_task)
                mem_avaiable = float(server['MEM_max']) - float(server['MEM_total'])
                mem_per_task = float(server['MEM_matlab']) / float(server['num_matlab'])
                num_mem = math.floor(mem_avaiable/mem_per_task)
                num_target = min([num_cpu, num_mem])
            df_temp = df[df['HostName']=='']
            if len(df_temp) > 0:#still have some tasks to do
                intialTaskIdx = list(df_temp.index)
                random.shuffle(intialTaskIdx)
                if len(intialTaskIdx) > num_target:
                    intialTaskIdx = intialTaskIdx[:num_target]
                for i in range(len(intialTaskIdx)):
                    df.loc[intialTaskIdx[i], 'HostName'] = server['name']
        writeJSON_from_df(task_path, df)
    
    def generateTasks(self):
        self.lastModifiedTime = self.getFileLastModifiedTime()
        lastModifiedTimeStr = datetime.strftime(self.lastModifiedTime, "%Y%m%d_%H%M%S")
        oldJsonName = os.path.join(self.finishedTaskFolder, 'TaskList_'+lastModifiedTimeStr+'.json')
        newJsonName = os.path.join(self.newTaskFolder, 'TaskList_'+lastModifiedTimeStr+'.json')
        if os.path.isfile(oldJsonName):# already finished
            return
        if os.path.isfile(newJsonName):# already created
            return
        parameterTable = pd.read_excel(self.taskFilePath, sheet_name='ParameterRange')
#        Helper.printTable(parameterTable, 'parameterTable')
        parameterIdxs = list(range(0,5))+list(range(7,15+5))
        parameterNames = list(parameterTable.columns)
        parameterNames_Target = [parameterNames[i] for i in parameterIdxs]
#        ServerNameList = parameterTable.loc[:, 'ServerName']
#        ServerNameList = [name for name in ServerNameList if type(name) == str]
#        ServerFactorList = parameterTable.loc[:, 'ServerFactor']
#        ServerFactorList = [factor for factor in ServerFactorList if not np.isnan(factor)]
#        ServerFactorArray = np.array(ServerFactorList)
#        ServerFactorArray = ServerFactorArray/sum(ServerFactorArray)
#        numServerName = len(ServerNameList)
        numTotalTask = 1;
        for name in parameterNames_Target:
            vector = parameterTable.loc[:, name]
            numNonNan = list(np.isnan(vector)).count(False)
            numTotalTask *= numNonNan
        
        columns = ['Num'] + parameterNames_Target
        vectorList = []
        for name in parameterNames_Target:
            vector = parameterTable.loc[:, name]
            nonNanIdx = np.isnan(vector)
            nonNanIdx = [not(i) for i in nonNanIdx]
            vector = [float(vector[i]) for i in range(len(vector)) if nonNanIdx[i]]
            vectorList.append(vector)
        
        combinations = list(prod(*vectorList))
        combinations = [[i+1]+list(combinations[i]) for i in range(len(combinations))]
        taskTable = pd.DataFrame(data=combinations, columns=columns)
#        numTasks = len(taskTable)
#        numTaskPerServer = math.floor(numTasks/numServerName)
#        numLastServer = numTasks - (numServerName-1)*numTaskPerServer
        
#        ServerTasks = [];
#        numAssignedTasks = 0
#        for i in range(numServerName-1):
#            numTaskThisServer = int(round(numTasks*ServerFactorArray[i]))
#            numAssignedTasks += numTaskThisServer
#            ServerTasks += [ServerNameList[i]]*numTaskThisServer
#        
#        if numTasks > numAssignedTasks:
#            ServerTasks += [ServerNameList[-1]]*(numTasks-numAssignedTasks)
#        taskTable['ServerName'] = ServerTasks
        taskTable_old = pd.read_excel(self.taskFilePath, sheet_name = 'Sheet1')
        columns_old = list(taskTable_old.columns)
        areSameTasks = True
        for name in columns:
            if name != 'Num':
                vector_old = taskTable_old.loc[:, name]
                vector_old = sorted(set(vector_old))
                vector_new = parameterTable.loc[:, name]
                vector_new = sorted(vector_new)
                nonNanIdx = np.isnan(vector_new)
                nonNanIdx = [not(i) for i in nonNanIdx]
                vector_new = [vector_new[i] for i in range(len(vector_new)) if nonNanIdx[i]]
                if vector_new != vector_old:
                    areSameTasks = False
                    print(vector_new)
                    print(vector_old)
        
        taskTable['UUID'] = ''
        taskTable.loc[:, 'UUID'] = taskTable.loc[:, 'UUID'].apply(getUUID)
        book = load_workbook(self.taskFilePath)
        writer = pd.ExcelWriter(self.taskFilePath, engine='openpyxl')
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        
        if not areSameTasks:
            taskTable_new = pd.DataFrame(columns=columns_old)
            taskTable_new = pd.concat([taskTable, taskTable_new], axis=1)
#            for column in taskTable.columns:
#                print(len(taskTable_old.loc[:numTasks, column]))
#                print(taskTable.loc[:, column])
#                taskTable_new.loc[:numTasks, column] = taskTable.loc[:, column]
#            taskTable_new.loc[list(range(numTasks, len(columns_old))), :] = ''
            self.emptySheetExcludeHeaders(writer.book['Sheet1'])
            taskTable_new.to_excel(writer, sheet_name='Sheet1', startrow=1, header=False,index=False)
            writer.save()
        
        taskTable = pd.read_excel(self.taskFilePath, sheet_name = 'Sheet1')
        
        newXlsxName = os.path.join(self.mainFolder, 'Output', 'TaskList_'+lastModifiedTimeStr+'.xlsx')
        writeJSON_from_df(newJsonName, taskTable)
        shutil.copyfile(self.taskFilePath, newXlsxName)

if __name__ == '__main__':
    pass
