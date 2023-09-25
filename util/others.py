#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Jun  4 13:41:42 2021

@author: frank
"""
import os
import random
import time
import uuid
import sys
import pandas as pd
import psutil
from openpyxl import load_workbook
import importlib
import tabulate as tb
from PyTaskDistributor.util.extract import extract_after


def print_table(table):
    if type(table) == pd.DataFrame:
        msg = tb.tabulate(table.values, table.columns, tablefmt="grid")
    else:
        msg = tb.tabulate(table, tablefmt="grid")
    return msg + '\n'

def getLibTrackingObj():
    libNames = ['master', 'server']
    importedLibs = {}
    for name in libNames:
        importedLibs[name] = {
            'path': './PyTaskDistributor/core/{}.py'.format(name),
            'lastModifiedTime': 0,
            'obj': importlib.import_module(
                'PyTaskDistributor.core.{}'.format(name))
        }

    for name, lib in importedLibs.items():
        importedLibs[name]['lastModifiedTime'] =\
            os.path.getmtime(importedLibs[name]['path'])
    return importedLibs


def importLatestLib(importedLibs):
    for name, lib in importedLibs.items():
        newModifiedTime = os.path.getmtime(importedLibs[name]['path'])
        if importedLibs[name]['lastModifiedTime'] != newModifiedTime:
            importlib.reload(importedLibs[name]['obj'])
            print('Module ' + ucword(name) + ' is reloaded')
            importedLibs[name]['lastModifiedTime'] = newModifiedTime
    return importedLibs


def ucword(word=None):
    if word is not None:
        length = len(word)
        if length > 0:
            temp = word[0].upper()
            temp += word[1:length].lower()
            return temp
    return word


def get_uuid(_):
    rd = random.Random()
    #    rd.seed(0)
    uuid_str = str(uuid.UUID(int=rd.getrandbits(128)))
    return uuid_str[:5]


def make_dirs(folder):
    if not os.path.isdir(folder):
        os.makedirs(folder)


def delete_file(path):
    if os.path.isfile(path):
        os.unlink(path)


def sleep_mins(num_min):  # for KeyboardInterrupt
    for i in range(round(num_min * 60)):
        time.sleep(1)


def empty_sheet_exclude_headers(ws):
    for i, row in enumerate(ws.iter_rows()):
        if i != 0:
            for cell in row:
                cell.value = None


def get_latest_file_in_folder(folder, ending=''):
    if not os.path.isdir(folder):
        return None
    file_list = os.listdir(folder)
    path_list = [os.path.join(folder, file)
                 for file in file_list if file.endswith(ending)]
    path_list_file = list(filter(os.path.isfile, path_list))
    if len(path_list_file) == 0:
        return None
    mat_modified_time = [os.path.getmtime(path) for path in path_list_file]
    target_id = mat_modified_time.index(max(mat_modified_time))
    return path_list_file[target_id]


def get_latest_timestamp(folder, ending=''):
    if not os.path.isdir(folder):
        return None
    file_list = os.listdir(folder)
    path_list = [os.path.join(folder, file)
                 for file in file_list if file.endswith(ending)]
    path_list_file = list(filter(os.path.isfile, path_list))
    if len(path_list_file) == 0:
        return None
    mat_modified_time = [os.path.getmtime(path) for path in path_list_file]
    return max(mat_modified_time)


def wait_changes(args):
    watching_check = ['-w=' in o for o in args]
    if any(watching_check):
        current_time = time.time()
        watching_id = watching_check.index(True)
        watching_folder = extract_after(sys.argv[watching_id], '-w=.')
        print('Waiting max 30mins for the change of folder .{}'.format(
            watching_folder))
        watching_folder = watching_folder.split(os.path.sep)
        watching_folder = os.path.join(os.getcwd(), *watching_folder)
        time_interval_mins = 10/60  # check every 10 seconds
        loop_guard = round(30/time_interval_mins)
        while loop_guard > 0:
            loop_guard -= 1
            latest_time = get_latest_timestamp(watching_folder, '.py')
            if latest_time < current_time:
                sleep_mins(time_interval_mins)
            else:
                latest_file = get_latest_file_in_folder(watching_folder, 'py')
                print('{} is changed'.format(os.path.basename(latest_file)))
                break


def get_file_suffix(path):
    basename = os.path.basename(path)
    dot_location = [i for i in range(len(basename)) if basename[i] == '.']
    if len(dot_location) > 0:
        suffix = basename[dot_location[-1]:]
    else:
        suffix = ''
    return suffix


def update_xlsx_file(path_xlsx, df, sheet_name='Sheet1'):
    if not os.path.isfile(path_xlsx):
        df.to_excel(path_xlsx, index=False)
    else:
        writer = pd.ExcelWriter(path_xlsx, engine='openpyxl',
                                mode='a', if_sheet_exists='replace')
        writer.workbook = load_workbook(path_xlsx)
        worksheet = writer.sheets[sheet_name]
        empty_sheet_exclude_headers(worksheet)
        df2 = df.reset_index(drop=True)
        for row_index, row in df2.iterrows():
            for col_index in range(len(df2.columns)):
                column = df2.columns[col_index]
                if row[column] != "":
                    worksheet.cell(row_index+2, col_index+1).value = \
                        row[column]

        writer.close()


def mark_finished_tasks(task_table, results, task_id):
    for i in range(len(task_id)):
        idx = task_id[i]
        result = results[i]
        for key in result.keys():
            task_table.loc[idx, key] = result[key]
    return task_table


def get_process_list():
    data = []
    columns = ['pid', 'User', 'Name', 'CPU', 'Mem', 'Command']
    for proc in psutil.process_iter():
        try:
            # Get process name & pid from process object.
            #            cpu_percent = proc.cpu_percent(interval=0.01)
            #            print(cpu_percent)
            cmdline = ' '.join(proc.cmdline())
            data.append([proc.pid, proc.username(), proc.name(),
                         0, proc.memory_percent(), cmdline])
        except (psutil.NoSuchProcess, psutil.AccessDenied,
                psutil.ZombieProcess):
            pass
    return pd.DataFrame(data=data, columns=columns)


def get_num_processor():
    return psutil.cpu_count()


def is_zombie_process(pid):
    try:
        return psutil.Process(pid).status().lower() == 'zombie'
    except Exception:
        return True


def get_process_cpu(pid):
    try:
        return psutil.Process(pid).cpu_percent(interval=1)
    except Exception:
        return 0.0


def get_process_mem(pid):
    try:
        return psutil.Process(pid).memory_percent()
    except Exception:
        return 0.0


def send_email(subject, content, receiver="j.chen-2@tudelft.nl",
               image_path=None):
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.image import MIMEImage
    from email.mime.text import MIMEText
    import os
    sender_email = os.environ['senderEmail']
    sender_pass = os.environ['senderPass']
    server = smtplib.SMTP('smtp.gmail.com', 587)

    # Next, log in to the server
    server.starttls()
    server.login(sender_email, sender_pass)

    # Send the mail
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = receiver
    msg['Subject'] = subject

    body = content
    msg.attach(MIMEText(body, 'plain'))
    if image_path is not None:
        img_data = open(image_path, 'rb').read()
        image = MIMEImage(img_data, name=os.path.basename(image_path))
        msg.attach(image)
    text = msg.as_string()
    server.sendmail(sender_email, receiver, text)
